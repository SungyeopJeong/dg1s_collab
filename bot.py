# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import os
import datetime
from datetime import timedelta
from pytz import timezone, utc
import openpyxl
import requests
from bs4 import BeautifulSoup

application=Flask(__name__)

KST=timezone('Asia/Seoul')

@application.route('/colstid', methods=['POST'])
def after_stid(): # 학번 입력 후
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"] # 사용자 고유 키
    stid=req["action"]["detailParams"]["student_id"]["value"] # 벌점 부여할 학번
    #isstaff=False
    isstaff=True
    print(userid,stid)
    
    '''fr=open("/home/ubuntu/dg1s_collab/staff_data.txt","r") # staff_data와 비교
    lines=fr.readlines()
    fr.close()
    for line in lines:
        if userid==line.rstrip("\n") : isstaff=True'''
    if isstaff==False: # 생교부원이 아니다
        print("where2")
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "권한이 없습니다."  
                        }
                    }
                ]
            }
        }
    else:
        quickReplies=[] # 경고/벌점을 바로가기 응답 형태로 제공
        msgtxt=["경고 1회 추가","벌점 1점 추가"]
        for msg in msgtxt:
            quickReplies.append({ "action": "block",
                                  "label": msg[:2],
                                  "messageText": msg,
                                  "blockId": "",
                                  "extra": { "stid": stid, "type": msg[:2] }})
        res={
            "version": "2.0",
            "template": {
                "output": [
                    {
                        "simpleText": {
                            "text": "welcome"  
                        }
                    }
                ],
                "quickReplies": quickReplies
            }
        }
    return jsonify(res)

@application.route('/coltype', methods=['POST'])
def after_type(): # 유형 선택 후
    
    print(req["intent"]["id"])
    req=request.get_json() # 파라미터 값 불러오기
    stid=req["action"]["clientExtra"]["stid"] # 부여할 학
    typei=req["action"]["clientExtra"]["type"] # 선택한 유형
    print(stid, typei)
                                 
    quickReplies=[] # 사유를 바로가기 응답 형태로 제공
    msgtxt=["미소등","책상 미정리","의자 미정리","콘센트","캐리어"]
    for msg in msgtxt:
        quickReplies.append({ "action": "block",
                              "label": msg,
                              "messageText": "사유 : "+msg,
                              "blockId": "",
                              "extra": { "stid": stid, "type": typei, "reason": msg }})
    res={
        "version": "2.0",
        "template": {
            "quickReplies": quickReplies
        }
    }
    return jsonify(res)

@application.route('/colreason', methods=['POST'])
def after_reason(): # 사유 선택 후
    
    print(req["intent"]["id"])
    req=request.get_json() # 파라미터 값 불러오기
    stid=req["action"]["clientExtra"]["stid"] # 부여할 학번
    typei=req["action"]["clientExtra"]["type"] # 선택한 유형
    reason=req["action"]["clientExtra"]["reason"] # 선택한 사유
    printmsg=""
    print(stid,typei,reason)
                                 
    fr=open("/home/ubuntu/dg1s_collab/student_data.txt","r") # student_data에 업데이트
    lines=fr.readlines()
    fr.close()
    for line in lines:
        data=line.split(' ')
        datastid=data[0]
        if stid==datastid:
            datawarning=data[1]
            datapenalty=data[2]
            printmsg=stid+"\n이전 : 경고 "+datawarning+"회, 벌점 "+datapenalty+"점\n"
            if typei=="경고": datawarning=str(int(datawarning)+1)
            elif typei=="벌점": datapenalty=str(int(datapenalty)+1)
            printmsg+="이후 : 경고 "+datawarning+"회, 벌점 "+datapenalty+"점\n사유 : "+reason
            break

    now=datetime.datetime.utcnow()
    print(utc.localize(now).astimezone(KST))
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "simpleText": {
                        "text": printmsg
                    }
                }
            ]
        }
    }
    return jsonify(res)

@application.route('/colload', methods=['POST'])
def load_data(): # 경고/벌점 확인
     
    req=request.get_json() # 파라미터 값 불러오기

    res={
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "simpleText": {
                        "text": "hi"
                    }
                }
            ]
        }
    }
    return jsonify(res)
'''  
@application.route('/stid', methods=['POST'])
def input_stid(): # 학번 입력 함수
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid=req["action"]["detailParams"]["student_id"]["value"]
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # userdata 저장 및 변경
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]
        if dusid==userid: fw.write(userid+" "+stid+" 7 none 0 none none\n")
        else : fw.write(line)
    fw.close()
    res={
        "version": "2.0",
        "template": { "outputs": [ { "simpleText": { "text": "학번이 "+stid+"(으)로 등록되었습니다." } } ] }
    }
    return jsonify(res)

@application.route('/save', methods=['POST'])
def final_save(): # 최종 저장 함수
        
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # 좌석 저장 후 초기화
    lines=fr.readlines()
    fr.close()
    rw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    fw=open("/home/ubuntu/dg1s_bot/final save.txt","a")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0]; dstid=datas[1]; dday=datas[2]; dmeal=datas[3]
        dseat=datas[4]; dp1=datas[5]; dp2=datas[6].rstrip('\n')
        if dusid==userid:
            if dmeal=="아침": dmeal='0'
            elif dmeal=="점심": dmeal='1'
            elif dmeal=="저녁": dmeal='2'
            fw.write(dstid+" "+dday+" "+dmeal+" "+dseat+" -\n")
            if dp1!="none": fw.write(dp1+" "+dday+" "+dmeal+" "+dseat+" *\n")
            if dp2!="none": fw.write(dp2+" "+dday+" "+dmeal+" "+dseat+" *\n")
            rw.write(userid+" "+dstid+" 7 none 0 none none\n")
        else : rw.write(line) 
    rw.close()
    fw.close()
    
    res={
        "version": "2.0",
        "template": { "outputs": [ { "simpleText": { "text": "저장되었습니다." } } ] }
    }
    return jsonify(res)

@application.route('/reset', methods=['POST'])
def reset(): # 초기화
    
    now=datetime.datetime.utcnow() # Meal 계산
    Day=int(utc.localize(now).astimezone(KST).strftime("%w"))
    hour=int(utc.localize(now).astimezone(KST).strftime("%H"))
    minu=int(utc.localize(now).astimezone(KST).strftime("%M"))
    if (hour==6 and minu>=50) or (hour>=7 and hour<12) or (hour==12 and minu<30): Meal="아침" # 가장 최근 식사가 언제인지 자동 계산
    elif (hour==12 and minu>=30) or (hour>=13 and hour<18) or (hour==18 and minu<30): Meal="점심"
    else: 
        Meal="저녁"
        if (hour==6 and minu<50) or hour<=5 : Day=(Day+6)%7
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"]
    stid="none"
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r") # 초기화
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        datas=line.split(" ")
        dusid=datas[0];
        if dusid==userid: stid=datas[1]
        if dusid!=userid: fw.write(line)
    fw.write(userid+" "+stid+" 7 none 0 none none\n")
    fw.close()
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "carousel": {
                        "type": "basicCard",
                        "items": [
                            {
                                "title": "[저장 확인]",
                                "description": "학번    "+stid+"\n날짜    "+Days[Day]+"\n식사    "+Meal+"\n좌석    0",
                                "buttons": [
                                    { "action": "message", "label": "확인", "messageText": "확인했습니다." },
                                    { "action": "message", "label": "초기화", "messageText": "초기화" }
                                ]
                            },
                            { 
                                "thumbnail":{
                                    "imageUrl": "http://k.kakaocdn.net/dn/L689z/btqJ78BkcF5/oG7PgVEcPhCqma4ZwyvwAk/img_l.jpg", "fixedRatio": "true"
                                }
                            }
                        ]
                    }
                }
            ]
        }
    }
    return jsonify(res)

@application.route('/excel', methods=['POST'])
def to_excel(): # 엑셀 파일로 생성
    
    wb = openpyxl.load_workbook('Gbob.xlsx',data_only=True) # 엑셀 기본 형식
    
    fr=open("/home/ubuntu/dg1s_bot/final save.txt","r") # 엑셀 채워 넣기
    lines=fr.readlines()
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(" ")
        if len(datas)!=5: continue
        dstid=datas[0]; dday=int(datas[1]); dmeal=int(datas[2]); dseat=datas[3]
        col=dday*3+dmeal; row=int(dstid[2:])+3 
        if dseat==".": dseat="X"
        if 4<=col and col<=16:
            sheet=wb[dstid[:2]]
            sheet.cell(row,col).value=dseat
    fr.close()
    
    sh = wb['통계']
    j = 0
    for sheet in wb:
        if sheet.title not in classn: continue
        T = sheet.title; N = str(classN[j]+3)
        # 통계 칸 채우기
        sh.cell(2,6).value = int(lines[0].rstrip('\n'))
        sh.cell(j+2,5).value = int(N)-3
        sh.cell(j+2,4).value = "=COUNTA("+T+"!D4:P"+N+")/('통계'!$F$2*("+N+"-3))"
        sh.cell(j+2,4).number_format = "0.00%"
        for k in range(4,4+classN[j]):
            # 참여율 칸 채우기
            K = str(k)
            sheet.cell(k,17).value = "=COUNTA(D"+K+":P"+K+")/'통계'!$F$2"
            sheet.cell(k,17).number_format = "0%"
        j += 1
    
    wb.save("bob.xlsx")
    res={
        "version": "2.0",
        "template": {
            "outputs": [ { "simpleText": { "text": "Excel 파일 생성 완료" } } ]
        }
    }
    return jsonify(res)

@application.route('/upst', methods=['POST'])
def update_stid(): # 학번 갱신 함수
    
    updatestr=""
    # 형식: "이전 학번_새 학번_..." ex) "1301 2106 1316 2417" 
    
    fr=open("/home/ubuntu/dg1s_bot/user data.txt","r")
    lines=fr.readlines()
    fr.close()
    fw=open("/home/ubuntu/dg1s_bot/user data.txt","w")
    for line in lines:
        former_stid=line.split(" ")[1]
        i=updatestr.find(former_stid)
        if i!=-1: 
            new_stid=updatestr[i+5:i+9]
            line=line.replace(former_stid,new_stid)
        fw.write(line)
    fw.close()
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [ { "simpleText": { "text": "학번 갱신 완료" } } ]
        }
    }
    return jsonify(res)
'''
@application.route('/')
def index():
    return render_template("index.html")

filename=""
'''
@application.route('/texteditor')
def text_editor(): # 원하는 파일 사이트에서 보여주고 편집
    global filename
    filename=request.args.get('filename')
    fr=open("/home/ubuntu/dg1s_bot/"+filename+".txt","r")
    data_send=fr.readlines()
    fr.close()
    if filename=="user data": data_send.sort(key=lambda x:x[13:17]) # 학번 순 정렬
    return render_template("texteditor.html",data=data_send,name=filename)

@application.route('/filesave', methods=['GET','POST'])
def save_as_file(): # txt file 저장하기
    if request.method=='POST':
        fr=open("/home/ubuntu/dg1s_bot/"+filename+".txt","r")
        before=fr.read()
        fr.close()
        
        text=request.form['content']
        text=str(text)
        with open("/home/ubuntu/dg1s_bot/"+filename+".txt","w",encoding='utf-8') as f:
            f.write(text)
    
        now=datetime.datetime.utcnow()
        hour=utc.localize(now).astimezone(KST).strftime("%H")
        minu=utc.localize(now).astimezone(KST).strftime("%M")
        date=utc.localize(now).astimezone(KST).strftime("%d")
        month=utc.localize(now).astimezone(KST).strftime("%m")
        year=utc.localize(now).astimezone(KST).strftime("%Y")
        fw=open("/home/ubuntu/dg1s_bot/log.txt","a")
        fw.write('['+year+'-'+month+'-'+date+' '+hour+':'+minu+"] '"+filename+".txt' saved (Below is the contents before saving.)\n")
        fw.write(before+'\n')
        fw.close()
        
        return render_template("saved.html")
  
@application.route('/xlsave', methods=['GET','POST'])
def save_as_xlfile(): # file 저장하기
    if request.method == 'POST':
        f=request.files['xlfile']
        f.save("/home/ubuntu/dg1s_bot/"+secure_filename(f.filename))
        return render_template("saved.html")
  
@application.route('/dnldfile', methods=['GET','POST'])
def download_file(): # file 다운받기
    if request.method == 'POST':
        filename=request.form['downloadfilename']
        return send_file("/home/ubuntu/dg1s_bot/"+filename, attachment_filename=filename, as_attachment=True)

@application.route('/file')
def upload_n_download():
    files=os.listdir("/home/ubuntu/dg1s_bot")
    folders=[]
    for file in files:
        if not '.' in file: folders.append(file)
    for folder in folders:
        files.remove(folder)
    return render_template("file.html", files=files)
'''
@application.route('/status')
def record_status():
    index=int(request.args.get('index'))
    n=classN[index]
    
    stid=[]
    for i in range(1,classN[index]+1):
        id=classn[index]
        if i<10: id+='0'
        id+=str(i)
        stid.append(id)
    
    name=Name[index]
    
    record=[]
    for i in range(25):
        record.append([])
        for j in range(13):
            record[i].append('')
        record[i].append(0)
    fr=open("/home/ubuntu/dg1s_bot/final save.txt", "r")
    lines=fr.readlines()
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(' '); id=datas[0]; day=int(datas[1]); meal=int(datas[2]); seat=datas[3]
        if id[:2]==classn[index]:
            if 3*day+meal-4<0 or 3*day+meal-4>12: continue
            if record[int(id[2:4])-1][3*day+meal-4]=='': record[int(id[2:4])-1][13]+=1
            if seat==".": seat="X"
            record[int(id[2:4])-1][3*day+meal-4]=seat
    fr.close()
    mealN=int(lines[0].rstrip('\n'))
    for i in range(n):
        record[i][13]=str(round((record[i][13]/mealN)*100))+'%'
    
    return render_template("status.html", n=n, stid=stid, name=name, record=record)

if __name__ == "__main__":
    application.run(host='0.0.0.0', port=5000)
