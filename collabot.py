# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, render_template, send_file
from werkzeug.utils import secure_filename
import os
import datetime
from datetime import timedelta
from pytz import timezone, utc
import openpyxl
import requests
#from bs4 import BeautifulSoup
import random

application=Flask(__name__)

KST=timezone('Asia/Seoul')

@application.route('/colstid', methods=['POST'])
def after_stid(): # 학번 입력 후
    
    req=request.get_json() # 파라미터 값 불러오기
    userid=req["userRequest"]["user"]["properties"]["plusfriendUserKey"] # 사용자 고유 키
    stid=req["action"]["detailParams"]["student_id"]["value"] # 벌점 부여할 학번
    isstaff=False
    staff=""
    
    #print(userid,stid)
    
    fr=open("/home/ubuntu/dg1s_collab/staff_data.txt","r") # staff_data와 비교
    lines=fr.readlines()
    fr.close()
    for line in lines:
        if userid in line: 
            isstaff=True
            staff=line.split(' ')[0]+' '+line.split(' ')[1] # staff 학번과 이름
    
    if isstaff==False: # 생교부원이 아니다
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": "권한이 없습니다." 
                        }
                    }
                ]
            }
        }
    else:
        quickReplies=[] # 경고/벌점을 바로가기 응답 형태로 제공
        msgtxt=["경고 1회 추가","벌점 1점 추가"]
        for msg in msgtxt:
            quickReplies.append({ "action": "block",
                                  "label": msg[:2],
                                  "messageText": msg,
                                  "blockId": "60c3a762a0293f369849360a",
                                  "extra": { "staff": staff, "stid": stid, "type": msg[:2] }})
        res={
            "version": "2.0",
            "template": {
                "outputs":[
                    {
                        "simpleText": {
                            "text": "경고/벌점을 선택해주세요."
                        }
                    }
                ],
                "quickReplies": quickReplies
            }
        }
    return jsonify(res)

@application.route('/coltype', methods=['POST'])
def after_type(): # 유형 선택 후
    
    req=request.get_json() # 파라미터 값 불러오기
    staff=req["action"]["clientExtra"]["staff"] # 생교부원 학번 이름
    stid=req["action"]["clientExtra"]["stid"] # 부여할 학번
    typei=req["action"]["clientExtra"]["type"] # 선택한 유형
                                 
    quickReplies=[] # 사유를 바로가기 응답 형태로 제공
    if typei=="경고": msgtxt=["미소등","책상_미정리","의자_미정리","콘센트","기타"]
    elif typei=="벌점": msgtxt=["캐리어","30분_이후 통행","타학생_책상에_두고_미정리"]
    for msg in msgtxt:
        if msg=="기타":
            quickReplies.append({ "action": "block",
                              "label": msg,
                              "messageText": "사유 : "+msg,
                              "blockId": "60c4a96bc13fe4037f226823",
                              "extra": { "staff": staff, "stid": stid, "type": typei }})
        else :
            msgforprint=msg.replace('_',' ')
            quickReplies.append({ "action": "block",
                              "label": msgforprint,
                              "messageText": "사유 : "+msgforprint,
                              "blockId": "60c3a77bcb976d4f0ad40ffa",
                              "extra": { "staff": staff, "stid": stid, "type": typei, "reason": msg }})
    res={
        "version": "2.0",
        "template": {
            "outputs":[
                {
                    "simpleText": {
                        "text": "사유를 선택해주세요."
                    }
                }
            ],
            "quickReplies": quickReplies
        }
    }
    return jsonify(res)

@application.route('/colask',methods=['POST'])
def ask_etc_reason():

    req=request.get_json() # 파라미터 값 불러오기
    staff=req["action"]["clientExtra"]["staff"] # 생교부원 학번 이름    
    stid=req["action"]["clientExtra"]["stid"] # 부여할 학번
    typei=req["action"]["clientExtra"]["type"] # 선택한 유형
    
    res={ # etc_reason OUTPUT context에 parameter 추가
        "version": "2.0",
        "context": {
            "values": [
                {
                    "name": "etc_reason",
                    "lifeSpan": 1,
                    "params": {
                        "checked": "true",
                        "staff": staff,
                        "stid": stid,
                        "type": typei
                    }
                }
            ]
        },
        "template": {
            "outputs": [
                {
                    "simpleText": {
                        "text": "기타 사유를 입력해주세요."
                    }
                }
            ]
        }
    }
    return jsonify(res)

def give_wp(staff,stid,typei,reason):
    
    now=datetime.datetime.utcnow() # 현재 시간
    time='['+utc.localize(now).astimezone(KST).strftime("%Y-%m-%d %H:%M:%S")+']'
    
    printmsg="" # 출력용 메시지
    logmsg="" # log 기록용 메시지
                                 
    fr=open("/home/ubuntu/dg1s_collab/student_data.txt","r") # student_data 불러와서
    backup=fr.read()
    lines=backup.split("\n")
    fr.close()
    
    fw=open("/home/ubuntu/dg1s_collab/backup.txt","a") # 혹시 모르니 백업
    fw.write(time+' '+staff+"의 기록 이전 데이터\n")
    fw.write(backup+"\n")
    fw.close()
    
    fw=open("/home/ubuntu/dg1s_collab/student_data.txt","w") # student_data 수정
    fw2=open("/home/ubuntu/dg1s_collab/log.txt","a") # log 기록
    for line in lines:
        data=line.rstrip("\n").split(' ')
        if len(data)<4: continue
        datastid=data[0]
        datawarning=data[1]
        datapenalty=data[2]
        datareason=data[3:]
        if stid==datastid:
            printmsg="[부여 완료]\n대상 : "+stid+"\n경고 "+datawarning+"회, 벌점 "+datapenalty+"점\n"
            logmsg=stid+" : "+datawarning+' '+datapenalty+' > '
            
            if typei=="경고": datawarning=str(int(datawarning)+1) # 경고 추가
            elif typei=="벌점": datapenalty=str(int(datapenalty)+1) # 벌점 추가
            if datawarning=='3': # 경고 3회면 벌점 1점으로
                datawarning='0'
                datapenalty=str(int(datapenalty)+1)
            datareason.append(time[1:11]+reason)
            
            printmsg+="> 경고 "+datawarning+"회, 벌점 "+datapenalty+"점\n사유 : "+reason.replace('_',' ')
            logmsg+=datawarning+' '+datapenalty+' '+reason
            fw2.write(time+' '+staff+", "+logmsg+"\n")
        fw.write(datastid+' '+datawarning+' '+datapenalty+' '+' '.join(datareason)+"\n")
    fw.close()
    fw2.close()
    return printmsg

@application.route('/colreason', methods=['POST'])
def after_reason(): # 사유 선택 후
        
    req=request.get_json() # 파라미터 값 불러오기
    staff=req["action"]["clientExtra"]["staff"] # 생교부원 학번 이름    
    stid=req["action"]["clientExtra"]["stid"] # 부여할 학번
    typei=req["action"]["clientExtra"]["type"] # 선택한 유형
    reason=req["action"]["clientExtra"]["reason"] # 선택한 사유
    printmsg=give_wp(staff,stid,typei,reason)
    
    res={
        "version": "2.0",
        "template": {
            "outputs": [
                {
                    "simpleText": {
                        "text": printmsg
                    }
                }
            ]
        }
    }
    return jsonify(res)

@application.route('/colfall',methods=['POST'])
def fall_back():
    
    req=request.get_json() # 파라미터 값 불러오기
    utter=req["userRequest"]["utterance"] # 입력한 내용
    checked=req["action"]["detailParams"]["checked"]["value"] # "true" > 기타 사유 입력 / "false" > 폴백
    
    if checked=="true":
        staff=req["action"]["detailParams"]["staff"]["value"] # 생교부원 학번 이름    
        stid=req["action"]["detailParams"]["stid"]["value"] # 부여할 학번
        typei=req["action"]["detailParams"]["type"]["value"] # 선택한 유형
        printmsg=give_wp(staff,stid,typei,utter.replace(' ','_'))
        
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": printmsg
                        }
                    }
                ]
            }
        }
    elif checked=="false":
        himsg=["나도 안녕","ㅎㅇㅎㅇ","안녕","하이"]
        laughmsg=["ㅋㅋㅋ","ㅋㅋㅋㅋ","ㅋㅋㅋㅋㅋ","ㅋㅋㅋㅋㅋㅋ","ㅎㅎㅎ"]
        sorrymsg=["미안","미안해","미안..."]
        randommsg=["아 알지(사실 모름)","나는 시키는 말밖에 못해","\n\n.___________.","?","음...?"]
        if "안녕" in utter or "hi" in utter or "ㅎㅇ" in utter:
            fallbackmsg=himsg
        elif "ㅋ" in utter or "ㅎ" in utter:
            fallbackmsg=laughmsg
        elif "아니" in utter or "ㅇㄴ" in utter or "답답" in utter:
            fallbackmsg=sorrymsg
        else :
            fallbackmsg=randommsg
        res={
            "version": "2.0",
            "template": {
                "outputs": [
                    {
                        "simpleText": {
                            "text": random.choice(fallbackmsg)
                        }
                    }
                ]
            }
        }
    return jsonify(res)

@application.route('/colstdata')
def show_stdata():
    
    fr=open("/home/ubuntu/dg1s_collab/student_data.txt","r") # student_data 불러와서
    data=fr.read()
    fr.close()
    
    return data

@application.route('/')
def main():
    
    gc=request.args.get('gc')
    if gc not in ["11","12","13","14","21","22","23","24","31","32","33","34"]: gc="11"
    stid=[]
    height=[]
    titles=[]
    
    fr=open("/home/ubuntu/dg1s_collab/student_data.txt","r")
    lines=fr.readlines()
    fr.close()

    for i in range(21):
        if i<9: stid.append(gc+'0'+str(i+1))
        else : stid.append(gc+str(i+1))
            
        index=(int(gc[0])-1)*4+int(gc[1])-1
        line=lines[index*21+i].rstrip('\n').split(' ')
        height.append(line[1])
        height.append(line[2])
        
        warning=line[1]
        penalty=line[2]
        reasons=line[3:]
        title="경고 "+warning+"회, 벌점 "+penalty+"점"
        if len(reasons)!=1:
            reasonmsg=""
            for reason in reasons:
                if reason=="none": continue
                reasonmsg+="\n"+reason.replace('_',' ')[:10]+' '+reason.replace('_',' ')[10:]
            title+="\n사유 :"+reasonmsg
        titles.append(title)
    
    return render_template("main.html",stid=stid,height=height,titles=titles)

@application.route('/excel')
def to_excel(): # 엑셀 파일로 생성
    
    wb = openpyxl.load_workbook('경고 및 벌점 표.xlsx',data_only=True) # 엑셀 기본 형식
    
    fr=open("/home/ubuntu/dg1s_collab/student_data.txt","r") # 엑셀 채워 넣기
    lines=fr.readlines()
    fr.close()
    for line in lines:
        data=line.rstrip('\n').split(' ')
        if len(data)<4: continue
        datastid=data[0]
        datawarning=data[1]
        datapenalty=data[2]
        datareason=data[3:]
        if stid==datastid:
            printmsg="[경고/벌점 현황]\n학번 : "+stid+"\n경고 "+datawarning+"회, 벌점 "+datapenalty+"점"
            if len(datareason)!=1:
                reasons=""
                for reason in datareason:
                    if reason=="none": continue
                    reasons+="\n"+reason.replace('_',' ')[:10]+' '+reason.replace('_',' ')[10:]
                printmsg+="\n사유 :"+reasons
    
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(" ")
        if len(datas)!=5: continue
        dstid=datas[0]; dday=int(datas[1]); dmeal=int(datas[2]); dseat=datas[3]
        col=dday*3+dmeal; row=int(dstid[2:])+3 
        if dseat==".": dseat="X"
        if 4<=col and col<=16:
            sheet=wb[dstid[:2]]
            sheet.cell(row,col).value=dseat
    
    sh = wb['통계']
    j = 0
    for sheet in wb:
        if sheet.title not in classn: continue
        T = sheet.title; N = str(classN[j]+3)
        # 통계 칸 채우기
        sh.cell(2,6).value = int(lines[0].rstrip('\n'))
        sh.cell(j+2,5).value = int(N)-3
        sh.cell(j+2,4).value = "=COUNTA("+T+"!D4:P"+N+")/('통계'!$F$2*("+N+"-3))"
        sh.cell(j+2,4).number_format = "0.00%"
        for k in range(4,4+classN[j]):
            # 참여율 칸 채우기
            K = str(k)
            sheet.cell(k,17).value = "=COUNTA(D"+K+":P"+K+")/'통계'!$F$2"
            sheet.cell(k,17).number_format = "0%"
        j += 1
    
    wb.save("bob.xlsx")
    res={
        "version": "2.0",
        "template": {
            "outputs": [ { "simpleText": { "text": "Excel 파일 생성 완료" } } ]
        }
    }
    return jsonify(res)

if __name__ == "__main__":
    application.run(host='0.0.0.0', port=5000)
'''

filename=""

@application.route('/texteditor')
def text_editor(): # 원하는 파일 사이트에서 보여주고 편집
    global filename
    filename=request.args.get('filename')
    fr=open("/home/ubuntu/dg1s_bot/"+filename+".txt","r")
    data_send=fr.readlines()
    fr.close()
    if filename=="user data": data_send.sort(key=lambda x:x[13:17]) # 학번 순 정렬
    return render_template("texteditor.html",data=data_send,name=filename)

@application.route('/filesave', methods=['GET','POST'])
def save_as_file(): # txt file 저장하기
    if request.method=='POST':
        fr=open("/home/ubuntu/dg1s_bot/"+filename+".txt","r")
        before=fr.read()
        fr.close()
        
        text=request.form['content']
        text=str(text)
        with open("/home/ubuntu/dg1s_bot/"+filename+".txt","w",encoding='utf-8') as f:
            f.write(text)
    
        now=datetime.datetime.utcnow()
        hour=utc.localize(now).astimezone(KST).strftime("%H")
        minu=utc.localize(now).astimezone(KST).strftime("%M")
        date=utc.localize(now).astimezone(KST).strftime("%d")
        month=utc.localize(now).astimezone(KST).strftime("%m")
        year=utc.localize(now).astimezone(KST).strftime("%Y")
        fw=open("/home/ubuntu/dg1s_bot/log.txt","a")
        fw.write('['+year+'-'+month+'-'+date+' '+hour+':'+minu+"] '"+filename+".txt' saved (Below is the contents before saving.)\n")
        fw.write(before+'\n')
        fw.close()
        
        return render_template("saved.html")
  
@application.route('/xlsave', methods=['GET','POST'])
def save_as_xlfile(): # file 저장하기
    if request.method == 'POST':
        f=request.files['xlfile']
        f.save("/home/ubuntu/dg1s_bot/"+secure_filename(f.filename))
        return render_template("saved.html")
  
@application.route('/dnldfile', methods=['GET','POST'])
def download_file(): # file 다운받기
    if request.method == 'POST':
        filename=request.form['downloadfilename']
        return send_file("/home/ubuntu/dg1s_bot/"+filename, attachment_filename=filename, as_attachment=True)

@application.route('/file')
def upload_n_download():
    files=os.listdir("/home/ubuntu/dg1s_bot")
    folders=[]
    for file in files:
        if not '.' in file: folders.append(file)
    for folder in folders:
        files.remove(folder)
    return render_template("file.html", files=files)

@application.route('/status')
def record_status():
    index=int(request.args.get('index'))
    n=classN[index]
    
    stid=[]
    for i in range(1,classN[index]+1):
        id=classn[index]
        if i<10: id+='0'
        id+=str(i)
        stid.append(id)
    
    name=Name[index]
    
    record=[]
    for i in range(25):
        record.append([])
        for j in range(13):
            record[i].append('')
        record[i].append(0)
    fr=open("/home/ubuntu/dg1s_bot/final save.txt", "r")
    lines=fr.readlines()
    for line in lines:
        if line==lines[0]: continue
        if "none" in line: continue
        datas=line.split(' '); id=datas[0]; day=int(datas[1]); meal=int(datas[2]); seat=datas[3]
        if id[:2]==classn[index]:
            if 3*day+meal-4<0 or 3*day+meal-4>12: continue
            if record[int(id[2:4])-1][3*day+meal-4]=='': record[int(id[2:4])-1][13]+=1
            if seat==".": seat="X"
            record[int(id[2:4])-1][3*day+meal-4]=seat
    fr.close()
    mealN=int(lines[0].rstrip('\n'))
    for i in range(n):
        record[i][13]=str(round((record[i][13]/mealN)*100))+'%'
    
    return render_template("status.html", n=n, stid=stid, name=name, record=record)
'''
